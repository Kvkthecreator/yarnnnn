"""Spreadsheet handler — table spec → XLSX via openpyxl."""

import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


async def render_spreadsheet(input_data: dict, output_format: str) -> tuple[bytes, str]:
    """
    Render a table spec to XLSX.

    input_data: {
        "title": str,
        "sheets": [
            {
                "name": str,
                "headers": [str, ...],
                "rows": [[value, ...], ...]
            },
            ...
        ]
    }
    output_format: "xlsx"
    Returns: (file_bytes, content_type)
    """
    if output_format != "xlsx":
        raise ValueError(f"Unsupported spreadsheet format: {output_format}")

    sheets = input_data.get("sheets", [])
    if not sheets:
        # Single sheet from flat data
        sheets = [{
            "name": input_data.get("title", "Sheet1"),
            "headers": input_data.get("headers", []),
            "rows": input_data.get("rows", []),
        }]

    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")

    for sheet_data in sheets:
        ws = wb.create_sheet(title=sheet_data.get("name", "Sheet"))
        headers = sheet_data.get("headers", [])
        rows = sheet_data.get("rows", [])

        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Write data rows
        for row_idx, row in enumerate(rows, 2):
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Auto-width columns
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
