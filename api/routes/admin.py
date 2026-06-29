"""
Admin routes — Operational dashboard for yarnnn.

Endpoints:
- GET /stats - Overview: users, agents, tasks, sessions
- GET /token-usage - Token & cost analytics from agent_runs + session_messages
- GET /execution-stats - Task execution frequency, credits, scheduler health
- GET /users - User list with activity metrics
- GET /export/users - Export users as Excel
- GET /export/report - Export full report as Excel
- POST /trigger-agent/{agent_id} - Manually trigger agent run (testing)
"""

import os
import logging
from io import BytesIO
from datetime import datetime, timedelta, timezone
from collections import defaultdict
from typing import Optional

from fastapi import APIRouter, HTTPException, Header
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

from services.admin_auth import AdminAuth

logger = logging.getLogger(__name__)
router = APIRouter()

# ADR-291: Use _BILLING_RATES from services.telemetry as single source of truth.
# Admin dashboard uses Anthropic cost (not user-facing markup) for internal analytics.
# Anthropic rates = user-facing rates / 2 (2x markup constant).
from services.telemetry import _BILLING_RATES

_ANTHROPIC_RATES = {
    model: {"input": r["input_per_mtok"] / 2 / 1000, "output": r["output_per_mtok"] / 2 / 1000}
    for model, r in _BILLING_RATES.items()
}
_DEFAULT_ANTHROPIC_RATE = _ANTHROPIC_RATES.get("claude-sonnet-4-6", {"input": 0.003, "output": 0.015})


# =============================================================================
# Pydantic Models
# =============================================================================

class AdminOverviewStats(BaseModel):
    total_users: int
    total_agents: int
    total_tasks: int
    total_sessions: int
    total_messages: int
    # Growth (7d)
    users_7d: int
    tasks_7d: int
    sessions_7d: int


class TokenUsageRow(BaseModel):
    """Per-day token usage breakdown.

    billed_input_tokens: fresh input tokens charged at full rate (excludes cache_read).
    cache_read_tokens: tokens served from cache at ~10% rate.
    cache_creation_tokens: tokens written to cache (slightly above input rate).
    The dashboard previously showed input_tokens from _total_input_tokens() which
    summed all three — making runs look 2-3x more expensive than they were. Now
    separated so cost and volume are independently legible.
    """
    date: str
    caller: str  # "chat", "task_pipeline", "other"
    model: str
    billed_input_tokens: int   # fresh input only — excludes cache_read
    output_tokens: int
    cache_read_tokens: int
    cache_creation_tokens: int
    api_calls: int
    estimated_cost_usd: float  # accurate: billed_input + cache_read*0.1 + cache_create*1.25 + output


class AdminTokenUsage(BaseModel):
    """Token usage analytics."""
    period_days: int
    total_billed_input_tokens: int
    total_output_tokens: int
    total_cache_read_tokens: int
    total_cache_creation_tokens: int
    total_api_calls: int
    total_estimated_cost_usd: float
    cache_hit_pct: float
    by_day: list[TokenUsageRow]


class TaskExecutionRow(BaseModel):
    """Per-task execution stats."""
    task_slug: str
    agent_title: str
    agent_role: str
    runs_total: int
    runs_7d: int
    avg_input_tokens: int
    avg_output_tokens: int
    last_run_at: Optional[str]
    # ADR-250 Phase 4 — from execution_events
    cost_usd_total: Optional[float] = None
    failed_count: int = 0
    skipped_count: int = 0


class AdminExecutionStats(BaseModel):
    """Task execution & scheduler health."""
    # Execution summary
    total_runs_24h: int
    total_runs_7d: int
    total_runs_30d: int
    # Spend
    spend_usd_this_month: float
    spend_usd_limit: float
    daily_spend_today: float = 0.0
    daily_spend_ceiling: float = 10.0
    # Scheduler
    last_scheduler_heartbeat: Optional[str]
    heartbeats_24h: int
    # Per-task breakdown
    tasks: list[TaskExecutionRow]


class AdminUserRow(BaseModel):
    id: str
    email: str
    created_at: str
    tier: str
    agent_count: int
    task_count: int
    session_count: int
    spend_usd: float
    last_activity: Optional[str] = None


class AdminAccountRow(BaseModel):
    """Per-persona live health for the designated test/eval accounts.

    Joins docs/alpha/personas.yaml (the test-account registry) against the
    execution_events / workspace_file_versions ledgers. Substrate-derived,
    no eval-specific schema — see docs/evaluations/README.md for the markdown
    side of the eval discipline this surface complements.
    """
    slug: str
    label: Optional[str] = None
    program: Optional[str] = None
    email: Optional[str] = None
    user_id: str
    # Wake activity (execution_events)
    wakes_24h: int
    wakes_7d: int
    failed_7d: int
    top_failure_reason: Optional[str] = None
    cost_7d: float
    last_wake: Optional[str] = None
    # Tenure signal (workspace_file_versions, authored_by reviewer:*)
    reviewer_edits_7d: int


# --- /accounts/{slug} detail blocks ---------------------------------------

class AccountDayPoint(BaseModel):
    day: str
    wakes: int
    cost: float
    input_tokens: int
    output_tokens: int


class AccountSlugRow(BaseModel):
    slug: str
    runs: int
    failed: int
    cost: float


class AccountSourceRow(BaseModel):
    wake_source: str
    success: int
    failed: int


class AccountFailureRow(BaseModel):
    created_at: str
    slug: str
    error_reason: Optional[str] = None
    error_detail: Optional[str] = None


class AccountRevisionRow(BaseModel):
    created_at: str
    path: str
    message: Optional[str] = None


class AccountProposalSummary(BaseModel):
    status: str
    count: int


class AccountPerfRow(BaseModel):
    mode: str
    avg_envelope_ms: Optional[int] = None
    avg_duration_ms: Optional[int] = None
    avg_tool_rounds: Optional[float] = None
    n: int


class AdminAccountDetail(BaseModel):
    """Full per-persona forensic view for /admin/accounts/{slug}.

    Seven blocks, all sliced from execution_events / workspace_file_versions /
    action_proposals / workspace_files. Substrate-derived; no eval-specific
    table. The daily curve covers up to `curve_days`, but high-frequency
    personas may truncate at the row cap — `curve_truncated` flags that so a
    cut tail reads as cut, not as zero.
    """
    slug: str
    label: Optional[str] = None
    program: Optional[str] = None
    email: Optional[str] = None
    user_id: str
    # Block 1: daily cost/activity curve
    curve_days: int
    curve_truncated: bool
    daily: list[AccountDayPoint]
    # Block 2: per-recurrence breakdown (7d)
    by_slug: list[AccountSlugRow]
    # Block 3: wake-source x status (7d)
    by_source: list[AccountSourceRow]
    # Block 4: recent failures with detail
    recent_failures: list[AccountFailureRow]
    # Block 5: reviewer self-amendment trail
    reviewer_trail: list[AccountRevisionRow]
    # Block 6: decision queue (proposals)
    proposals: list[AccountProposalSummary]
    # Block 7: substrate footprint + perf
    total_files: int
    persona_files: int
    operation_files: int
    perf: list[AccountPerfRow]


# =============================================================================
# Helper
# =============================================================================

def _get_date_threshold(days: int) -> str:
    return (datetime.now(timezone.utc) - timedelta(days=days)).isoformat()


def _estimate_cost(
    model: str,
    billed_input_tokens: int,
    output_tokens: int,
    cache_read_tokens: int = 0,
    cache_creation_tokens: int = 0,
) -> float:
    """Estimate Anthropic cost (not user-facing rate) from token counts and model.

    Cache pricing (Anthropic rates as of 2026-04):
      cache_read:     10% of input rate
      cache_creation: 125% of input rate
    This gives an accurate cost figure rather than treating all input as full-price.
    """
    pricing = _ANTHROPIC_RATES.get(model, _DEFAULT_ANTHROPIC_RATE)
    input_rate = pricing["input"]
    output_rate = pricing["output"]
    cost = (
        (billed_input_tokens / 1000 * input_rate)
        + (cache_read_tokens / 1000 * input_rate * 0.10)
        + (cache_creation_tokens / 1000 * input_rate * 1.25)
        + (output_tokens / 1000 * output_rate)
    )
    return cost


# =============================================================================
# GET /stats — Overview
# =============================================================================

@router.get("/stats", response_model=AdminOverviewStats)
async def get_overview_stats(admin: AdminAuth):
    """Overview dashboard statistics."""
    try:
        client = admin.client
        seven_days_ago = _get_date_threshold(7)

        # Total users
        users_result = client.table("workspaces").select("id", count="exact").execute()
        total_users = users_result.count or 0

        users_7d_result = client.table("workspaces")\
            .select("id", count="exact")\
            .gte("created_at", seven_days_ago).execute()
        users_7d = users_7d_result.count or 0

        # Total agents
        agents_result = client.table("agents").select("id", count="exact").execute()
        total_agents = agents_result.count or 0

        # Total tasks
        tasks_result = client.table("tasks").select("id", count="exact").execute()
        total_tasks = tasks_result.count or 0

        tasks_7d_result = client.table("tasks")\
            .select("id", count="exact")\
            .gte("created_at", seven_days_ago).execute()
        tasks_7d = tasks_7d_result.count or 0

        # Sessions
        sessions_result = client.table("chat_sessions").select("id", count="exact").execute()
        total_sessions = sessions_result.count or 0

        sessions_7d_result = client.table("chat_sessions")\
            .select("id", count="exact")\
            .gte("created_at", seven_days_ago).execute()
        sessions_7d = sessions_7d_result.count or 0

        # Messages
        messages_result = client.table("session_messages").select("id", count="exact").execute()
        total_messages = messages_result.count or 0

        return AdminOverviewStats(
            total_users=total_users,
            total_agents=total_agents,
            total_tasks=total_tasks,
            total_sessions=total_sessions,
            total_messages=total_messages,
            users_7d=users_7d,
            tasks_7d=tasks_7d,
            sessions_7d=sessions_7d,
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to fetch stats: {str(e)}")


# =============================================================================
# GET /token-usage — Token & Cost Analytics
# =============================================================================

@router.get("/token-usage", response_model=AdminTokenUsage)
async def get_token_usage(admin: AdminAuth, days: int = 7):
    """Token usage analytics from agent_runs and session_messages metadata."""
    try:
        client = admin.client
        cutoff = _get_date_threshold(min(days, 90))  # Cap at 90 days

        # --- Agent runs (task pipeline + legacy execution) ---
        runs_result = client.table("agent_runs")\
            .select("created_at, metadata")\
            .gte("created_at", cutoff)\
            .order("created_at", desc=True)\
            .limit(5000)\
            .execute()

        # --- Chat messages (TP + chat agents) ---
        # Only assistant messages have token metadata
        msgs_result = client.table("session_messages")\
            .select("created_at, metadata")\
            .eq("role", "assistant")\
            .gte("created_at", cutoff)\
            .not_.is_("metadata", "null")\
            .order("created_at", desc=True)\
            .limit(5000)\
            .execute()

        # Aggregate by day × caller
        # billed_input_tokens = fresh input only (excludes cache_read).
        # Separated so the dashboard shows real cost drivers, not inflated totals.
        daily: dict[str, dict[str, dict]] = defaultdict(lambda: defaultdict(lambda: {
            "billed_input_tokens": 0, "output_tokens": 0,
            "cache_read_tokens": 0, "cache_creation_tokens": 0,
            "api_calls": 0, "model": "",
        }))

        totals = {
            "billed_input_tokens": 0, "output_tokens": 0,
            "cache_read_tokens": 0, "cache_creation_tokens": 0,
            "api_calls": 0, "cost": 0.0,
        }

        # Process agent_runs → caller = "task_pipeline"
        for run in (runs_result.data or []):
            meta = run.get("metadata") or {}
            input_t = meta.get("input_tokens", 0) or 0
            output_t = meta.get("output_tokens", 0) or 0
            cache_read = meta.get("cache_read_input_tokens", 0) or 0
            cache_create = meta.get("cache_creation_input_tokens", 0) or 0
            model = meta.get("model", "claude-sonnet-4-6")

            date_str = run["created_at"][:10]
            caller = "task_pipeline"

            bucket = daily[date_str][caller]
            bucket["billed_input_tokens"] += input_t
            bucket["output_tokens"] += output_t
            bucket["cache_read_tokens"] += cache_read
            bucket["cache_creation_tokens"] += cache_create
            bucket["api_calls"] += 1
            bucket["model"] = model

            cost = _estimate_cost(model, input_t, output_t, cache_read, cache_create)
            totals["billed_input_tokens"] += input_t
            totals["output_tokens"] += output_t
            totals["cache_read_tokens"] += cache_read
            totals["cache_creation_tokens"] += cache_create
            totals["api_calls"] += 1
            totals["cost"] += cost

        # Process session_messages → caller = "chat"
        for msg in (msgs_result.data or []):
            meta = msg.get("metadata") or {}
            input_t = meta.get("input_tokens", 0) or 0
            output_t = meta.get("output_tokens", 0) or 0
            cache_read = meta.get("cache_read_input_tokens", 0) or 0
            cache_create = meta.get("cache_creation_input_tokens", 0) or 0
            model = meta.get("model", "claude-sonnet-4-6")

            if not input_t and not output_t:
                continue

            date_str = msg["created_at"][:10]
            caller = "chat"

            bucket = daily[date_str][caller]
            bucket["billed_input_tokens"] += input_t
            bucket["output_tokens"] += output_t
            bucket["cache_read_tokens"] += cache_read
            bucket["cache_creation_tokens"] += cache_create
            bucket["api_calls"] += 1
            bucket["model"] = model

            cost = _estimate_cost(model, input_t, output_t, cache_read, cache_create)
            totals["billed_input_tokens"] += input_t
            totals["output_tokens"] += output_t
            totals["cache_read_tokens"] += cache_read
            totals["cache_creation_tokens"] += cache_create
            totals["api_calls"] += 1
            totals["cost"] += cost

        # Build response rows
        by_day = []
        for date_str in sorted(daily.keys()):
            for caller, bucket in daily[date_str].items():
                model = bucket["model"] or "claude-sonnet-4-6"
                cost = _estimate_cost(
                    model,
                    bucket["billed_input_tokens"],
                    bucket["output_tokens"],
                    bucket["cache_read_tokens"],
                    bucket["cache_creation_tokens"],
                )
                by_day.append(TokenUsageRow(
                    date=date_str,
                    caller=caller,
                    model=model,
                    billed_input_tokens=bucket["billed_input_tokens"],
                    output_tokens=bucket["output_tokens"],
                    cache_read_tokens=bucket["cache_read_tokens"],
                    cache_creation_tokens=bucket["cache_creation_tokens"],
                    api_calls=bucket["api_calls"],
                    estimated_cost_usd=round(cost, 4),
                ))

        # Cache hit %: portion of all input-class tokens that were cache hits
        total_cacheable = totals["billed_input_tokens"] + totals["cache_read_tokens"] + totals["cache_creation_tokens"]
        cache_hit_pct = round(totals["cache_read_tokens"] / total_cacheable * 100, 1) if total_cacheable else 0.0

        return AdminTokenUsage(
            period_days=days,
            total_billed_input_tokens=totals["billed_input_tokens"],
            total_output_tokens=totals["output_tokens"],
            total_cache_read_tokens=totals["cache_read_tokens"],
            total_cache_creation_tokens=totals["cache_creation_tokens"],
            total_api_calls=totals["api_calls"],
            total_estimated_cost_usd=round(totals["cost"], 4),
            cache_hit_pct=cache_hit_pct,
            by_day=by_day,
        )

    except Exception as e:
        logger.error(f"[ADMIN] Token usage query failed: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to fetch token usage: {str(e)}")


# =============================================================================
# GET /execution-stats — Task Execution & Scheduler Health
# =============================================================================

@router.get("/execution-stats", response_model=AdminExecutionStats)
async def get_execution_stats(admin: AdminAuth):
    """Task execution frequency, credits, scheduler health."""
    try:
        client = admin.client
        now = datetime.now(timezone.utc)
        cutoff_24h = (now - timedelta(hours=24)).isoformat()
        cutoff_7d = (now - timedelta(days=7)).isoformat()
        cutoff_30d = (now - timedelta(days=30)).isoformat()

        # Run counts
        runs_24h = client.table("agent_runs").select("id", count="exact")\
            .gte("created_at", cutoff_24h).execute()
        runs_7d = client.table("agent_runs").select("id", count="exact")\
            .gte("created_at", cutoff_7d).execute()
        runs_30d = client.table("agent_runs").select("id", count="exact")\
            .gte("created_at", cutoff_30d).execute()

        # Spend this month (ADR-291: execution_events is canonical cost ledger)
        month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0).isoformat()
        spend_result = client.table("execution_events")\
            .select("cost_usd")\
            .gte("created_at", month_start)\
            .execute()
        spend_usd_this_month = sum(float(r.get("cost_usd") or 0) for r in (spend_result.data or []))

        # Spend limit (aggregate — $20 pro default shown for overall)
        spend_usd_limit = 20.0  # Pro default

        # Scheduler heartbeat
        hb_result = client.table("activity_log")\
            .select("created_at")\
            .eq("event_type", "scheduler_heartbeat")\
            .order("created_at", desc=True)\
            .limit(1).execute()
        last_heartbeat = hb_result.data[0]["created_at"] if hb_result.data else None

        hb_24h = client.table("activity_log")\
            .select("id", count="exact")\
            .eq("event_type", "scheduler_heartbeat")\
            .gte("created_at", cutoff_24h).execute()

        # Per-task breakdown from agent_runs
        # Get recent runs with agent + task metadata
        recent_runs = client.table("agent_runs")\
            .select("agent_id, created_at, metadata")\
            .gte("created_at", cutoff_30d)\
            .order("created_at", desc=True)\
            .limit(2000)\
            .execute()

        # Get agent info for display
        agents_result = client.table("agents")\
            .select("id, title, role")\
            .execute()
        agent_map = {a["id"]: a for a in (agents_result.data or [])}

        # Aggregate per task_slug
        task_stats: dict[str, dict] = {}
        for run in (recent_runs.data or []):
            meta = run.get("metadata") or {}
            task_slug = meta.get("task_slug", "unknown")
            agent_id = run["agent_id"]
            agent_info = agent_map.get(agent_id, {})

            if task_slug not in task_stats:
                task_stats[task_slug] = {
                    "agent_title": agent_info.get("title", "Unknown"),
                    "agent_role": agent_info.get("role", "unknown"),
                    "runs_total": 0,
                    "runs_7d": 0,
                    "input_tokens": [],
                    "output_tokens": [],
                    "last_run_at": None,
                }

            ts = task_stats[task_slug]
            ts["runs_total"] += 1
            if run["created_at"] >= cutoff_7d:
                ts["runs_7d"] += 1
            if not ts["last_run_at"] or run["created_at"] > ts["last_run_at"]:
                ts["last_run_at"] = run["created_at"]

            input_t = meta.get("input_tokens", 0) or 0
            output_t = meta.get("output_tokens", 0) or 0
            if input_t:
                ts["input_tokens"].append(input_t)
            if output_t:
                ts["output_tokens"].append(output_t)

        # ADR-250 Phase 4: enrich with execution_events (cost + failure counts per slug)
        today_utc = now.replace(hour=0, minute=0, second=0, microsecond=0).isoformat()
        daily_spend_today = 0.0
        ee_cost_map: dict[str, float] = {}
        ee_failed_map: dict[str, int] = {}
        ee_skipped_map: dict[str, int] = {}
        try:
            ee_result = client.table("execution_events")\
                .select("slug, status, cost_usd, created_at")\
                .gte("created_at", cutoff_30d)\
                .execute()
            for ee in (ee_result.data or []):
                s = ee.get("slug", "unknown")
                cost = float(ee.get("cost_usd") or 0)
                ee_cost_map[s] = ee_cost_map.get(s, 0.0) + cost
                if ee.get("status") == "failed":
                    ee_failed_map[s] = ee_failed_map.get(s, 0) + 1
                if ee.get("status") == "skipped":
                    ee_skipped_map[s] = ee_skipped_map.get(s, 0) + 1
                if ee.get("created_at", "") >= today_utc:
                    daily_spend_today += cost
        except Exception as e:
            logger.warning("[ADMIN] execution_events enrichment failed (non-fatal): %s", e)

        tasks = []
        for slug, ts in sorted(task_stats.items(), key=lambda x: x[1]["runs_total"], reverse=True):
            avg_in = int(sum(ts["input_tokens"]) / len(ts["input_tokens"])) if ts["input_tokens"] else 0
            avg_out = int(sum(ts["output_tokens"]) / len(ts["output_tokens"])) if ts["output_tokens"] else 0
            tasks.append(TaskExecutionRow(
                task_slug=slug,
                agent_title=ts["agent_title"],
                agent_role=ts["agent_role"],
                runs_total=ts["runs_total"],
                runs_7d=ts["runs_7d"],
                avg_input_tokens=avg_in,
                avg_output_tokens=avg_out,
                last_run_at=ts["last_run_at"],
                cost_usd_total=round(ee_cost_map.get(slug, 0.0), 4) or None,
                failed_count=ee_failed_map.get(slug, 0),
                skipped_count=ee_skipped_map.get(slug, 0),
            ))

        import os as _os
        return AdminExecutionStats(
            total_runs_24h=runs_24h.count or 0,
            total_runs_7d=runs_7d.count or 0,
            total_runs_30d=runs_30d.count or 0,
            spend_usd_this_month=spend_usd_this_month,
            spend_usd_limit=spend_usd_limit,
            daily_spend_today=round(daily_spend_today, 4),
            daily_spend_ceiling=float(_os.getenv("DAILY_SPEND_CEILING_USD", "10.0")),
            last_scheduler_heartbeat=last_heartbeat,
            heartbeats_24h=hb_24h.count or 0,
            tasks=tasks,
        )

    except Exception as e:
        logger.error(f"[ADMIN] Execution stats query failed: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to fetch execution stats: {str(e)}")


# =============================================================================
# GET /users — User List
# =============================================================================

@router.get("/users", response_model=list[AdminUserRow])
async def list_users(admin: AdminAuth):
    """List users with activity metrics."""
    try:
        client = admin.client
        month_start = datetime.now(timezone.utc).replace(
            day=1, hour=0, minute=0, second=0, microsecond=0
        ).isoformat()

        workspaces_result = client.table("workspaces")\
            .select("owner_id, owner_email, created_at")\
            .order("created_at", desc=True)\
            .limit(100)\
            .execute()

        if not workspaces_result.data:
            return []

        users = []
        for workspace in workspaces_result.data:
            user_id = workspace["owner_id"]
            email = workspace.get("owner_email") or "unknown"

            # Agent count
            agents = client.table("agents")\
                .select("id", count="exact")\
                .eq("user_id", user_id).execute()

            # Task count
            tasks = client.table("tasks")\
                .select("id", count="exact")\
                .eq("user_id", user_id).execute()

            # Session count + last activity
            sessions = client.table("chat_sessions")\
                .select("id, created_at", count="exact")\
                .eq("user_id", user_id)\
                .order("created_at", desc=True)\
                .limit(1).execute()

            last_activity = sessions.data[0].get("created_at") if sessions.data else None

            # Spend this month (ADR-291: execution_events is canonical cost ledger)
            spend = client.table("execution_events")\
                .select("cost_usd")\
                .eq("user_id", user_id)\
                .gte("created_at", month_start)\
                .execute()
            spend_usd = sum(float(r.get("cost_usd") or 0) for r in (spend.data or []))

            # Tier
            from services.platform_limits import get_user_tier
            tier = get_user_tier(client, user_id)

            users.append(AdminUserRow(
                id=user_id,
                email=email,
                created_at=workspace["created_at"],
                tier=tier,
                agent_count=agents.count or 0,
                task_count=tasks.count or 0,
                session_count=sessions.count or 0,
                spend_usd=spend_usd,
                last_activity=last_activity,
            ))

        return users

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to fetch users: {str(e)}")


# =============================================================================
# GET /accounts — Per-Persona Test-Account Health
# =============================================================================

def _load_personas() -> list[dict]:
    """Read the test-account registry (docs/alpha/personas.yaml).

    Single source of truth for who each eval persona is (slug, user_id,
    program). Hat-B artifact per CLAUDE.md — dev-only, never surfaced to real
    operators. Returns [] if the file is missing (e.g. a deploy that doesn't
    ship docs/), so the surface degrades to empty rather than 500ing.
    """
    import yaml

    # api/routes/admin.py -> repo root is two parents up from api/
    here = os.path.dirname(os.path.abspath(__file__))
    repo_root = os.path.abspath(os.path.join(here, "..", ".."))
    path = os.path.join(repo_root, "docs", "alpha", "personas.yaml")
    if not os.path.exists(path):
        logger.warning("[ADMIN] personas.yaml not found at %s — /accounts empty", path)
        return []
    with open(path) as f:
        data = yaml.safe_load(f) or {}
    return data.get("personas", []) or []


@router.get("/accounts", response_model=list[AdminAccountRow])
async def list_accounts(admin: AdminAuth):
    """Per-persona live health for the designated test/eval accounts.

    Joins personas.yaml against execution_events (wake activity, cost,
    failures) and workspace_file_versions (Reviewer self-amendment count, the
    tenure signal). All substrate-derived — no eval-specific table.
    """
    try:
        client = admin.client
        personas = _load_personas()
        if not personas:
            return []

        cutoff_24h = _get_date_threshold(1)
        cutoff_7d = _get_date_threshold(7)

        rows: list[AdminAccountRow] = []
        for p in personas:
            user_id = p.get("user_id")
            if not user_id:
                continue

            # Wake activity over the last 7d (one fetch, bucketed in Python).
            # Scoped to 7d, not 30d: high-frequency personas (the traders wake
            # ~670×/day) blow past any row cap over 30d, which would silently
            # undercount cost. 7d stays well under the 5000 cap, so every
            # bucketed figure below — including cost — is exact, not truncated.
            events = client.table("execution_events")\
                .select("status, error_reason, cost_usd, created_at")\
                .eq("user_id", user_id)\
                .gte("created_at", cutoff_7d)\
                .order("created_at", desc=True)\
                .limit(5000)\
                .execute()
            event_rows = events.data or []

            wakes_24h = sum(1 for e in event_rows if e["created_at"] >= cutoff_24h)
            wakes_7d = len(event_rows)
            failed_7d = sum(1 for e in event_rows if e.get("status") == "failed")
            # cost_usd is NULL on mechanical/skipped wakes — coerce (the bug that
            # 500'd /execution-stats; `or 0` not `, 0` per the present-but-null trap).
            cost_7d = sum(float(e.get("cost_usd") or 0) for e in event_rows)
            last_wake = event_rows[0]["created_at"] if event_rows else None

            # Most common failure reason in the window (operator triage signal).
            reason_counts: dict[str, int] = defaultdict(int)
            for e in event_rows:
                if e.get("status") == "failed" and e.get("error_reason"):
                    reason_counts[e["error_reason"]] += 1
            top_failure_reason = (
                max(reason_counts, key=reason_counts.get) if reason_counts else None
            )

            # Tenure signal: Reviewer self-amendments (authored_by reviewer:*) in 7d.
            reviewer_edits = client.table("workspace_file_versions")\
                .select("id", count="exact")\
                .eq("user_id", user_id)\
                .like("authored_by", "freddie:%")\
                .gte("created_at", cutoff_7d)\
                .execute()

            rows.append(AdminAccountRow(
                slug=p.get("slug", "unknown"),
                label=p.get("label"),
                program=str(p["program"]) if p.get("program") else None,
                email=p.get("email"),
                user_id=user_id,
                wakes_24h=wakes_24h,
                wakes_7d=wakes_7d,
                failed_7d=failed_7d,
                top_failure_reason=top_failure_reason,
                cost_7d=round(cost_7d, 4),
                last_wake=last_wake,
                reviewer_edits_7d=reviewer_edits.count or 0,
            ))

        # Most-active first (mirrors the validated query's ORDER BY wakes_7d DESC).
        rows.sort(key=lambda r: r.wakes_7d, reverse=True)
        return rows

    except Exception as e:
        logger.error("[ADMIN] Accounts query failed: %s", e)
        raise HTTPException(status_code=500, detail=f"Failed to fetch accounts: {str(e)}")


@router.get("/accounts/{slug}", response_model=AdminAccountDetail)
async def get_account_detail(slug: str, admin: AdminAuth):
    """Full forensic view for one test/eval persona.

    Seven blocks sliced from execution_events / workspace_file_versions /
    action_proposals / workspace_files. Same NULL-coercion and substrate-only
    discipline as /accounts.
    """
    try:
        client = admin.client
        persona = next((p for p in _load_personas() if p.get("slug") == slug), None)
        if not persona:
            raise HTTPException(status_code=404, detail=f"Unknown persona: {slug}")
        user_id = persona.get("user_id")
        if not user_id:
            raise HTTPException(status_code=404, detail=f"Persona {slug} has no user_id")

        CURVE_DAYS = 30
        CURVE_CAP = 5000
        cutoff_curve = _get_date_threshold(CURVE_DAYS)
        cutoff_7d = _get_date_threshold(7)

        # --- Block 1: daily curve (lightweight columns, capped) -------------
        curve_events = client.table("execution_events")\
            .select("created_at, cost_usd, input_tokens, output_tokens")\
            .eq("user_id", user_id)\
            .gte("created_at", cutoff_curve)\
            .order("created_at", desc=True)\
            .limit(CURVE_CAP)\
            .execute()
        curve_rows = curve_events.data or []
        curve_truncated = len(curve_rows) >= CURVE_CAP

        day_buckets: dict[str, dict] = defaultdict(
            lambda: {"wakes": 0, "cost": 0.0, "in": 0, "out": 0}
        )
        for e in curve_rows:
            day = e["created_at"][:10]
            b = day_buckets[day]
            b["wakes"] += 1
            b["cost"] += float(e.get("cost_usd") or 0)
            b["in"] += int(e.get("input_tokens") or 0)
            b["out"] += int(e.get("output_tokens") or 0)
        daily = [
            AccountDayPoint(
                day=day, wakes=b["wakes"], cost=round(b["cost"], 4),
                input_tokens=b["in"], output_tokens=b["out"],
            )
            for day, b in sorted(day_buckets.items())
        ]

        # --- Blocks 2/3: per-slug + per-source (7d full rows, bucketed) -----
        events_7d = client.table("execution_events")\
            .select("slug, status, error_reason, error_detail, cost_usd, created_at, wake_source")\
            .eq("user_id", user_id)\
            .gte("created_at", cutoff_7d)\
            .order("created_at", desc=True)\
            .limit(CURVE_CAP)\
            .execute()
        rows_7d = events_7d.data or []

        slug_buckets: dict[str, dict] = defaultdict(
            lambda: {"runs": 0, "failed": 0, "cost": 0.0}
        )
        source_buckets: dict[str, dict] = defaultdict(lambda: {"success": 0, "failed": 0})
        for e in rows_7d:
            s = slug_buckets[e.get("slug") or "—"]
            s["runs"] += 1
            s["cost"] += float(e.get("cost_usd") or 0)
            if e.get("status") == "failed":
                s["failed"] += 1
            src = source_buckets[e.get("wake_source") or "—"]
            if e.get("status") == "failed":
                src["failed"] += 1
            else:
                src["success"] += 1
        by_slug = sorted(
            [AccountSlugRow(slug=k, runs=v["runs"], failed=v["failed"], cost=round(v["cost"], 4))
             for k, v in slug_buckets.items()],
            key=lambda r: r.runs, reverse=True,
        )
        by_source = sorted(
            [AccountSourceRow(wake_source=k, success=v["success"], failed=v["failed"])
             for k, v in source_buckets.items()],
            key=lambda r: r.success + r.failed, reverse=True,
        )

        # --- Block 4: recent failures with detail ---------------------------
        fail_result = client.table("execution_events")\
            .select("created_at, slug, error_reason, error_detail")\
            .eq("user_id", user_id)\
            .eq("status", "failed")\
            .order("created_at", desc=True)\
            .limit(15)\
            .execute()
        recent_failures = [
            AccountFailureRow(
                created_at=r["created_at"], slug=r.get("slug") or "—",
                error_reason=r.get("error_reason"),
                error_detail=(r.get("error_detail") or "")[:200] or None,
            )
            for r in (fail_result.data or [])
        ]

        # --- Block 5: reviewer self-amendment trail -------------------------
        trail_result = client.table("workspace_file_versions")\
            .select("created_at, path, message")\
            .eq("user_id", user_id)\
            .like("authored_by", "freddie:%")\
            .order("created_at", desc=True)\
            .limit(15)\
            .execute()
        reviewer_trail = [
            AccountRevisionRow(
                created_at=r["created_at"], path=r["path"],
                message=(r.get("message") or "")[:120] or None,
            )
            for r in (trail_result.data or [])
        ]

        # --- Block 6: decision queue (proposals, status mix) ----------------
        prop_result = client.table("action_proposals")\
            .select("status")\
            .eq("user_id", user_id)\
            .gte("created_at", cutoff_curve)\
            .limit(2000)\
            .execute()
        prop_counts: dict[str, int] = defaultdict(int)
        for r in (prop_result.data or []):
            prop_counts[r.get("status") or "unknown"] += 1
        proposals = sorted(
            [AccountProposalSummary(status=k, count=v) for k, v in prop_counts.items()],
            key=lambda r: r.count, reverse=True,
        )

        # --- Block 7: substrate footprint + perf ----------------------------
        files_result = client.table("workspace_files")\
            .select("path")\
            .eq("user_id", user_id)\
            .limit(5000)\
            .execute()
        file_paths = [r["path"] for r in (files_result.data or [])]
        total_files = len(file_paths)
        persona_files = sum(1 for p in file_paths if p.startswith("/workspace/persona/"))
        operation_files = sum(1 for p in file_paths if p.startswith("/workspace/operation/"))

        # Perf averages bucketed by mode (from the 7d full-row fetch would lack
        # perf columns; re-fetch the perf columns over 7d, capped).
        perf_result = client.table("execution_events")\
            .select("mode, envelope_load_ms, duration_ms, tool_rounds")\
            .eq("user_id", user_id)\
            .gte("created_at", cutoff_7d)\
            .limit(CURVE_CAP)\
            .execute()
        perf_buckets: dict[str, dict] = defaultdict(
            lambda: {"env": [], "dur": [], "rounds": [], "n": 0}
        )
        for r in (perf_result.data or []):
            mode = r.get("mode")
            if not mode:
                continue
            pb = perf_buckets[mode]
            pb["n"] += 1
            if r.get("envelope_load_ms") is not None:
                pb["env"].append(r["envelope_load_ms"])
            if r.get("duration_ms") is not None:
                pb["dur"].append(r["duration_ms"])
            if r.get("tool_rounds") is not None:
                pb["rounds"].append(r["tool_rounds"])

        def _avg_int(xs):
            return round(sum(xs) / len(xs)) if xs else None

        def _avg_float(xs):
            return round(sum(xs) / len(xs), 1) if xs else None

        perf = [
            AccountPerfRow(
                mode=mode,
                avg_envelope_ms=_avg_int(pb["env"]),
                avg_duration_ms=_avg_int(pb["dur"]),
                avg_tool_rounds=_avg_float(pb["rounds"]),
                n=pb["n"],
            )
            for mode, pb in perf_buckets.items()
        ]

        return AdminAccountDetail(
            slug=slug,
            label=persona.get("label"),
            program=str(persona["program"]) if persona.get("program") else None,
            email=persona.get("email"),
            user_id=user_id,
            curve_days=CURVE_DAYS,
            curve_truncated=curve_truncated,
            daily=daily,
            by_slug=by_slug,
            by_source=by_source,
            recent_failures=recent_failures,
            reviewer_trail=reviewer_trail,
            proposals=proposals,
            total_files=total_files,
            persona_files=persona_files,
            operation_files=operation_files,
            perf=perf,
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error("[ADMIN] Account detail query failed (%s): %s", slug, e)
        raise HTTPException(status_code=500, detail=f"Failed to fetch account detail: {str(e)}")


# =============================================================================
# Export Endpoints (kept from legacy — still useful for IR)
# =============================================================================

@router.get("/export/users")
async def export_users_excel(admin: AdminAuth):
    """Export users data as Excel file."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        users = await list_users(admin)

        wb = Workbook()
        ws = wb.active
        ws.title = "Users"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

        headers = ["Email", "Tier", "Agents", "Tasks", "Sessions", "Spend (mo)", "Last Active", "Joined"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border

        for row, user in enumerate(users, 2):
            ws.cell(row=row, column=1, value=user.email).border = thin_border
            ws.cell(row=row, column=2, value=user.tier).border = thin_border
            ws.cell(row=row, column=3, value=user.agent_count).border = thin_border
            ws.cell(row=row, column=4, value=user.task_count).border = thin_border
            ws.cell(row=row, column=5, value=user.session_count).border = thin_border
            ws.cell(row=row, column=6, value=user.spend_usd).border = thin_border
            ws.cell(row=row, column=7, value=user.last_activity or "—").border = thin_border
            ws.cell(row=row, column=8, value=user.created_at).border = thin_border

        widths = [35, 8, 8, 8, 10, 10, 25, 25]
        for col, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
        ws.freeze_panes = "A2"

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=yarnnn_users_{timestamp}.xlsx"}
        )
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to export: {str(e)}")


@router.get("/export/report")
async def export_full_report(admin: AdminAuth):
    """Export comprehensive report as multi-sheet Excel."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter

        client = admin.client
        now = datetime.now(timezone.utc)

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        section_font = Font(bold=True, size=12, color="4F46E5")
        metric_label_font = Font(bold=True, size=11)
        metric_value_font = Font(size=14, bold=True)
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

        # Fetch data
        stats = await get_overview_stats(admin)
        token_usage = await get_token_usage(admin, days=30)
        exec_stats = await get_execution_stats(admin)
        users = await list_users(admin)

        wb = Workbook()

        # --- Summary Sheet ---
        ws = wb.active
        ws.title = "Summary"
        ws.merge_cells("A1:D1")
        ws["A1"].value = "yarnnn — Platform Metrics Report"
        ws["A1"].font = Font(bold=True, size=16)
        ws["A2"] = f"Generated: {now.strftime('%B %d, %Y at %H:%M UTC')}"
        ws["A2"].font = Font(italic=True, color="666666")

        row = 4
        ws[f"A{row}"] = "PLATFORM"
        ws[f"A{row}"].font = section_font
        row += 1
        for label, value in [
            ("Total Users", stats.total_users),
            ("Users (7d)", stats.users_7d),
            ("Total Agents", stats.total_agents),
            ("Total Tasks", stats.total_tasks),
            ("Total Sessions", stats.total_sessions),
            ("Total Messages", stats.total_messages),
        ]:
            ws[f"A{row}"] = label
            ws[f"A{row}"].font = metric_label_font
            ws[f"B{row}"] = value
            ws[f"B{row}"].font = metric_value_font
            row += 1

        row += 1
        ws[f"A{row}"] = "TOKEN COSTS (30d)"
        ws[f"A{row}"].font = section_font
        row += 1
        for label, value in [
            ("Total API Calls", token_usage.total_api_calls),
            ("Billed Input Tokens", f"{token_usage.total_billed_input_tokens:,}"),
            ("Cache Read Tokens", f"{token_usage.total_cache_read_tokens:,}"),
            ("Output Tokens", f"{token_usage.total_output_tokens:,}"),
            ("Estimated Cost", f"${token_usage.total_estimated_cost_usd:.2f}"),
            ("Cache Hit %", f"{token_usage.cache_hit_pct:.1f}%"),
        ]:
            ws[f"A{row}"] = label
            ws[f"A{row}"].font = metric_label_font
            ws[f"B{row}"] = value
            row += 1

        row += 1
        ws[f"A{row}"] = "EXECUTION (30d)"
        ws[f"A{row}"].font = section_font
        row += 1
        for label, value in [
            ("Task Runs (24h)", exec_stats.total_runs_24h),
            ("Task Runs (7d)", exec_stats.total_runs_7d),
            ("Task Runs (30d)", exec_stats.total_runs_30d),
            ("Spend USD (month)", exec_stats.spend_usd_this_month),
        ]:
            ws[f"A{row}"] = label
            ws[f"A{row}"].font = metric_label_font
            ws[f"B{row}"] = value
            row += 1

        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 20

        # --- Users Sheet ---
        ws_users = wb.create_sheet("Users")
        headers = ["Email", "Tier", "Agents", "Tasks", "Sessions", "Credits", "Last Active", "Joined"]
        for col, h in enumerate(headers, 1):
            cell = ws_users.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        for r, u in enumerate(users, 2):
            for c, v in enumerate([u.email, u.tier, u.agent_count, u.task_count,
                                   u.session_count, u.spend_usd, u.last_activity or "—", u.created_at], 1):
                ws_users.cell(row=r, column=c, value=v).border = thin_border
        for col, w in enumerate([35, 8, 8, 8, 10, 10, 25, 25], 1):
            ws_users.column_dimensions[get_column_letter(col)].width = w
        ws_users.freeze_panes = "A2"

        # --- Token Usage Sheet ---
        ws_tokens = wb.create_sheet("Token Usage")
        t_headers = ["Date", "Caller", "Model", "Billed Input", "Output Tokens", "Cache Read", "API Calls", "Cost ($)"]
        for col, h in enumerate(t_headers, 1):
            cell = ws_tokens.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        for r, row_data in enumerate(token_usage.by_day, 2):
            for c, v in enumerate([row_data.date, row_data.caller, row_data.model,
                                   row_data.billed_input_tokens, row_data.output_tokens,
                                   row_data.cache_read_tokens, row_data.api_calls,
                                   row_data.estimated_cost_usd], 1):
                ws_tokens.cell(row=r, column=c, value=v).border = thin_border
        ws_tokens.freeze_panes = "A2"

        # --- Task Executions Sheet ---
        ws_tasks = wb.create_sheet("Task Executions")
        e_headers = ["Task", "Agent", "Role", "Runs (30d)", "Runs (7d)", "Avg Input Tokens", "Avg Output Tokens", "Last Run"]
        for col, h in enumerate(e_headers, 1):
            cell = ws_tasks.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        for r, t in enumerate(exec_stats.tasks, 2):
            for c, v in enumerate([t.task_slug, t.agent_title, t.agent_role, t.runs_total,
                                   t.runs_7d, t.avg_input_tokens, t.avg_output_tokens,
                                   t.last_run_at or "—"], 1):
                ws_tasks.cell(row=r, column=c, value=v).border = thin_border
        ws_tasks.freeze_panes = "A2"

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        timestamp = now.strftime("%Y%m%d_%H%M%S")
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=yarnnn_report_{timestamp}.xlsx"}
        )
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to export: {str(e)}")


# =============================================================================
# Admin Testing Endpoints
# =============================================================================

@router.post("/trigger-task/{task_slug}")
async def admin_trigger_task(
    task_slug: str,
    x_service_key: Optional[str] = Header(None),
) -> dict:
    """Trigger a recurrence for testing.

    Per ADR-261 D3 + ADR-296 v2 D1: walks /workspace/_recurrences.yaml
    for the user that owns the slug and submits a wake proposal through
    the manual_fire wake source.
    """
    from services.wake_sources.manual_fire import fire as wake_manual_fire
    from services.recurrence import walk_workspace_recurrences

    supabase_key = os.environ.get("SUPABASE_SERVICE_KEY", "")
    if not x_service_key or x_service_key != supabase_key:
        raise HTTPException(status_code=403, detail="Invalid service key")

    supabase_url = os.environ.get("SUPABASE_URL", "")
    if not supabase_url or not supabase_key:
        raise HTTPException(status_code=500, detail="Missing Supabase credentials")

    from supabase import create_client
    client = create_client(supabase_url, supabase_key)

    task_result = client.table("tasks").select("user_id, slug").eq("slug", task_slug).limit(1).execute()
    if not task_result.data:
        raise HTTPException(status_code=404, detail=f"No task index row for slug: {task_slug}")

    user_id = task_result.data[0]["user_id"]
    recurrences = walk_workspace_recurrences(client, user_id)
    rec = next((r for r in recurrences if r.slug == task_slug), None)
    if rec is None:
        raise HTTPException(
            status_code=404,
            detail=f"No recurrence entry for slug '{task_slug}' (user {user_id[:8]}).",
        )

    try:
        # ADR-296 v2 D1: admin manual fire routes through the manual_fire
        # wake source. Funnel auto-escalates per D1 (operator explicit
        # assertion is wake-warrant).
        result = await wake_manual_fire(client, user_id, rec)
        return {
            "task_slug": task_slug,
            "success": result.get("success", False),
            "message": result.get("message"),
            "trigger": result.get("trigger"),
            "actions_taken": result.get("actions_taken"),
        }
    except Exception as e:
        import traceback
        return {"task_slug": task_slug, "status": "error", "error": str(e), "traceback": traceback.format_exc()}
